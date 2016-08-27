
import spacy
from spacy.en import English
import spacy.parts_of_speech
import json

nlp=English()

#question=nlp(u'what color is the sweater?')

gencount=0

#----------load VQA Validation data questions-------------------

data_file=open("data/OpenEnded_mscoco_val2014_questions.json","r")

data=json.load(data_file)

questiondata=[]

for entry in data['questions']:
	questiondata.append([entry['question'],0])

#---------------------------------------------------------------


#------------ Load General Knowledge Questions -----------------

import openpyxl

wb=openpyxl.load_workbook('genericqdata.xlsx')

sheets=wb.get_sheet_names()
#questiondata=[]
for sheet in sheets:
	sheet_name=wb.get_sheet_by_name(sheet)
	numdata=sheet_name.get_highest_row()
	for i in range(1,numdata):
		if sheet_name.cell(row=i,column=2).value!=None:
			questiondata.append([sheet_name.cell(row=i,column=2).value,1])

#---------------------------------------------------------------

#-----------load philosophy questions---------------------------
#
data_file=open("data/philosophicalquestions.txt")
data=data_file.readlines()
#questiondata=[]
for entry in data:
	questiondata.append([entry.decode('utf-8'),1])
#
#---------------------------------------------------------------

import random

random.seed(1)

random.shuffle(questiondata)

#test_questiondata=questiondata[100000:131464]

tmp=zip(*questiondata)
questions=tmp[0]
labels=[]
labels_sp=[]
for entry in tmp[1]:
	if entry!=None:
		labels.append(entry) # for class generic
		labels_sp.append(1-entry) #for class specific


pred_labels=[]
pred_base_label=[]

pred_labels_sp=[]
pred_base_label_sp=[]

qcount=0
alllens=[]
genlen=[]
splen=[]

for index,qentry in enumerate(questions[100001:131000]):
	if qentry!=None:
		qcount=qcount+1
		question=nlp(qentry)	

		tags=[]

		for word in question:
			tags.append([word,word.tag_,word.dep_])

		ind=0
		dind=0
		nind=0
		cvind=0
		g_ind=0
		ccind=0
		phil_ind=0
		for item in tags:

			if item[1]=='NN' or item[1]=='NNP':
			#	print "noun"
		
				if dind==1:
					ind=0
					#g_ind=0
				elif dind==0:
					if nind==1:
						ind=ind
					else:
						ind=1
						#g_ind=1
				nind=1
				dind=0
			#	print dind
				if cvind==1:
					ind=0
			else:
				
				nind=0
		
			if item[1]=='NNS':
				if dind==1:
					ind=0
					#g_ind=0
				elif dind==0:
					if nind==1:
						ind=ind
					else:
						ind=1
						#g_ind=1
				nind=1
				dind=0
			#	print dind
				if cvind==1:
					ind=0
			else:
				
				nind=0

			#print dind

			if item[1]=='EX':
		
				dind=1

			if item[1]=='VBG': #or item[1]=='VBN':
				cvind=1
				ind=0
	
			if item[1]=='DT':
				dind=1
				ind=0
	
			if item[1]=='VBZ':
				ind=1
	
			if item[1]=='PRP$' or item[1]=='POS' or item[1]=='PRP':
				#g_ind=0
				dind=1
				ind=0
	
			if str(item[0])== 'here' or str(item[0])=='there' or str(item[0])=='there ':
				g_ind=0
				ind=0

			if str(item[0])== 'this ' or str(item[0])=='these ' or str(item[0])=='they ' or str(item[0])=='shown' or str(item[0])=='seen' or str(item[0])=='visible' or str(item[0])=='see ':
				g_ind=0				
				ind=0
				cvind=1

			if str(item[0])=='life' or str(item[0])=='meaning ' or str(item[0])=='God ':
				g_ind=1

			
			if item[1]=='CC':
				if nind==1:
					nind=1
				else:
					nind=0
				ccind=1
	

	
		if qcount%10000==0:
			print qcount

		if ind==1 or g_ind==1:
			gencount=gencount+1
			pred_labels.append(1) #for class generic
			pred_labels_sp.append(0) # for class specific
#			print question
#			print "Generic"
#			print '\n'
		else:
			pred_labels.append(0) #for class generic
			pred_labels_sp.append(1) #for class specific
#			print question
#			print "Specific"
#			print '\n'
		
		if random.random()<0.08:
			pred_base_label.append(1)
			pred_base_label_sp.append(0)
		else:
			pred_base_label.append(0)
			pred_base_label_sp.append(1)
		

		#if len(tags)>8:
		#	pred_base_label.append(1)
		#	pred_base_label_sp.append(0)
		#else:
		#	pred_base_label.append(0)
		#	pred_base_label_sp.append(1)
		#alllens.append(len(tags))
		
		#if labels[50000+index]==1:
		#	genlen.append(len(tags))
		#else:
		#	splen.append(len(tags))



labelstst=labels[100001:131000]
labels_sptst=labels_sp[100001:131000]

from sklearn.metrics import precision_score
from sklearn.metrics import recall_score
from sklearn.metrics import accuracy_score
#========== ACCURACY =============================
a=accuracy_score(labelstst,pred_labels)
print "Accuracy : "+str(a)

#========== FOR CLASS GENERAL ====================
print "For Class General (our Algorithm)"
#Our Scores
p=precision_score(labelstst,pred_labels)
print "precision: "+ str(p)
r=recall_score(labelstst,pred_labels)
print "recall: "+ str(r)

#Baseline Scores
print "For Class General (biased coin baseline)"
p_b=precision_score(labelstst,pred_base_label)
print "Precision : "+ str(p_b)
r_b=recall_score(labelstst,pred_base_label)
print "Recall : "+ str(r_b)

#========== FOR CLASS SPECIFIC ===================
print "For class Specific (our algo)"
#Our Scores
p_sp=precision_score(labels_sptst,pred_labels_sp)
print "Precision : "+ str(p_sp)
r_sp=recall_score(labels_sptst,pred_labels_sp)
print "Recall : "+ str(r_sp)

print "For class Specific (biased coin baseline)"
#Baseline Scores
p_b_sp=precision_score(labels_sptst,pred_base_label_sp)
print "Precision : "+ str(p_b_sp)
r_b_sp=recall_score(labels_sptst,pred_base_label_sp)
print "Recall : "+ str(r_b_sp)

print "Normalized accuracy for our rule base algo : " + str((r+r_sp)/2)

