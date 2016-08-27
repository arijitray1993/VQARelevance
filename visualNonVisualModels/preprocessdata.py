import spacy
from spacy.en import English
import spacy.parts_of_speech
import json

nlp=English()

import numpy as np


gencount=0

#----------load VQA Validation data questions-------------------

data_file=open("data/OpenEnded_mscoco_val2014_questions.json","r")

data=json.load(data_file)

questiondata=[]

for entry in data['questions']:
	questiondata.append([entry['question'],0])

#------------ Load General Knowledge Questions -----------------

import openpyxl

wb=openpyxl.load_workbook('genericqdata.xlsx')

sheets=wb.get_sheet_names()
#questiondata=[]
for sheet in sheets:
	sheet_name=wb.get_sheet_by_name(sheet)
	numdata=sheet_name.get_highest_row()
	for i in range(1,numdata):
		if sheet_name.cell(row=i,column=2).value!=None:
			questiondata.append([sheet_name.cell(row=i,column=2).value,1])

#---------------------------------------------------------------

#-----------load philosophy questions---------------------------
#
data_file=open("data/philosophicalquestions.txt")
data=data_file.readlines()
#questiondata=[]
for entry in data:
	questiondata.append([entry.decode('utf-8'),1])
#
#---------------------------------------------------------------

import random

random.seed(1)

random.shuffle(questiondata)

#with open("non-visual_questions.txt","w") as f:
#	for entry in questiondata[9750:9850]:
#		print>>f,entry[0]


tmp=zip(*questiondata)
questions=tmp[0]
labels=[]
labels_sp=[]
for entry in tmp[1]:
	if entry!=None:
		labels.append(entry) # for class generic
		labels_sp.append(1-entry) #for class specific

question_tags=[]

unique_tags=['CC','CD', 'DT', 'EX' ,'FW', 'IN', 'JJ', 'JJR', 'JJS', 'LS', 'MD', 'NN', 'NNS', 'NNP', 'NNPS', 'PDT', 'POS', 'PRP', 'PRP$', 'RB', 'RBR', 'RBS', 'RP', 'SYM', 'TO', 'UH', 'VB', 'VBD', 'VBG', 'VBN','VBP','VBZ','WDT', 'WP', 'WP$', 'WRB', 'SP'] 

vocab_size=len(unique_tags)

for qentry in questions:
        if qentry!=None:
                #qcount=qcount+1
                question=nlp(qentry)

                tags=[]

                for word in question:
			tag=word.tag_
			if tag in unique_tags:
				index=unique_tags.index(tag)
                        	tags.append(index)

                question_tags.append(tags)



'''
i=1
for entry in question_tags:
	entry=np.asarray(entry)
	onehot=np.zeros((len(entry),len(unique_tags)))
	onehot[np.arange(len(entry)),entry] = 1
	
	tag_onehot[i]=onehot
	i=i+1

'''
'''
import csv
myfile = open('sentencedata.csv', 'wb')
wr = csv.writer(myfile)
wr.writerow(question_tags)
'''


